import sys

from flask import Flask, render_template, request, jsonify, redirect, url_for, Response, stream_with_context
import serial
import time
import os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from datetime import datetime
import webbrowser
import threading
import os
app = Flask(__name__)

# Get the correct base path whether running as .py or .exe
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

template_dir = os.path.join(BASE_DIR, 'templates')
static_dir = os.path.join(BASE_DIR, 'static')

app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)

# Global state to support two teams
state = {
    "teams_enabled": False,  # Toggle Team B on/off
    "active_team": "Team A",
    "teams": {
        "Team A": {
            "balls_counter": 0,
            "allRoundsData": []
        },
        "Team B": {
            "balls_counter": 0,
            "allRoundsData": []
        }
    },
    "current_round": 1
}



def save_points_to_excel(total, defects, balls_in_play, plan_value):
    if (total + defects + balls_in_play) > 100:
        unprocessed = 0
    else:
        unprocessed = 100 - (total + defects + balls_in_play)

    delta = (total - defects + 0.5 * balls_in_play) - plan_value
    points = (15 * total) if abs(delta) == 0 else (10 * total) / (abs(delta) / 9)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = os.path.join(os.path.dirname(__file__), "game_scores.xlsx")

    # Load or create workbook
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Total", "Defects", "Balls in Play", "Balls Unprocessed", "Points", "Timestamp"])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    # Append new data
    ws.append([total, defects, balls_in_play, unprocessed, round(points, 2), timestamp])

    # Convert rows (excluding header) to list for sorting
    data = list(ws.iter_rows(min_row=2, values_only=True))
    # Sort by points (index 4) in descending order
    sorted_data = sorted(data, key=lambda row: row[4], reverse=True)

    # Clear old data rows (but keep header)
    ws.delete_rows(2, ws.max_row)

    # Re-add sorted rows
    for row in sorted_data:
        ws.append(row)

    wb.save(filename)

def reset_arduino():
    try:
        arduino = serial.Serial(port='/dev/ttyACM0', baudrate=9600, timeout=0.5)  # Shorter timeout
        time.sleep(0.5)  # Shorter wait for Arduino to be ready
        arduino.write(b'r')  # Send reset command

        # Wait for confirmation for up to 0.5 seconds
        deadline = time.time() + 0.5
        while time.time() < deadline:
            response = arduino.readline().decode().strip()
            if "Counter reset to 0" in response:
                break

        # Immediately clear counter in state
        state["teams"][state["active_team"]]["balls_counter"] = 0
        arduino.close()
    except serial.SerialException as e:
        print(f"Error: {e}")


def get_arduino_connection():
    try:
        arduino = serial.Serial(port='/dev/ttyACM0', baudrate=9600, timeout=1)
        time.sleep(2)
        return arduino
    except Exception as e:
        print(f"Error connecting to Arduino: {e}")
        return None


@app.route('/')
@app.route('/plan')
def plan():
    leaderboard = []

    filename = "game_scores.xlsx"
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        data = list(ws.iter_rows(min_row=2, values_only=True))
        # Sort by points (column 4 = index 4), descending
        sorted_data = sorted(data, key=lambda row: row[4], reverse=True)
        # Take top 3
        leaderboard = sorted_data[:3]

    return render_template('plan.html',
                           current_round=state["current_round"],
                           active_team=state["active_team"],
                           teams_enabled=state["teams_enabled"],
                           leaderboard=leaderboard)

@app.route('/toggle_team', methods=['POST'])
def toggle_team():
    state["teams_enabled"] = not state["teams_enabled"]
    state["active_team"] = "Team A"
    return redirect(url_for('plan'))


@app.route('/switch_team', methods=['POST'])
def switch_team():
    if state["teams_enabled"]:
        state["active_team"] = "Team B" if state["active_team"] == "Team A" else "Team A"
    return jsonify({"active_team": state["active_team"]})


@app.route('/timer', methods=['GET', 'POST'])
def timer():
    if request.method == 'POST':
        plan_value = request.form.get('plan_value')
        if not plan_value or not plan_value.isdigit():
            return render_template('plan.html', current_round=state["current_round"], error="Invalid input",
                                   active_team=state["active_team"], teams_enabled=state["teams_enabled"])
        plan_value = int(plan_value)

        team_data = state["teams"][state["active_team"]]
        existing_round = next((r for r in team_data["allRoundsData"] if r["round"] == state["current_round"]), None)
        if existing_round:
            existing_round["plan"] = plan_value
        else:
            team_data["allRoundsData"].append({
                "round": state["current_round"],
                "plan": plan_value,
                "total": 0,
                "defects": 0
            })

    #  RESET THE ARDUINO HERE before rendering timer.html
    reset_arduino()

    balls = state["teams"][state["active_team"]]["balls_counter"]
    return render_template('timer.html', balls_counter=balls, current_round=state["current_round"],
                           active_team=state["active_team"])


@app.route('/update_counter', methods=['GET'])
def update_counter():
    arduino = get_arduino_connection()
    if arduino:
        try:
            response = arduino.readline().decode().strip()
            if response.isdigit():
                count = int(response)
                state["teams"][state["active_team"]]["balls_counter"] = count
                return jsonify({"balls_counter": count})
        except Exception as e:
            print(f"Error fetching live counter: {e}")
        finally:
            arduino.close()
    return jsonify({"error": "Failed to connect to Arduino"}), 500


@app.route('/get_round_data/<int:round_number>')
def get_round_data(round_number):
    team_data = state["teams"][state["active_team"]]
    round_data = next((r for r in team_data["allRoundsData"] if r["round"] == round_number), None)
    if round_data:
        actual = round_data["total"] - round_data["defects"]
        balls_in_play = round_data.get("balls_in_play", 0)
        delta = (actual + 0.5 * balls_in_play) - round_data["plan"]
        return jsonify({
            "plan": round_data["plan"],
            "total": round_data["total"],
            "defects": round_data["defects"],
            "actual": actual,
            "delta": delta
        })
    return jsonify({"error": "Round not found"}), 404


@app.route('/defects')
def defects():
    team_data = state["teams"][state["active_team"]]
    last_round = next((r for r in team_data["allRoundsData"] if r["round"] == state["current_round"]), None)
    defects = last_round["defects"] if last_round else 0
    return render_template('defects.html', current_round=state["current_round"], defects=defects,
                           active_team=state["active_team"])


@app.route('/update_defects', methods=['POST'])
def update_defects():
    defects_value = request.form.get('defects')
    if not defects_value or not defects_value.isdigit():
        return jsonify({"error": "Invalid defects value"}), 400

    defects_value = int(defects_value)
    team_data = state["teams"][state["active_team"]]
    for round_data in team_data["allRoundsData"]:
        if round_data["round"] == state["current_round"]:
            round_data["defects"] = defects_value
            break

    return jsonify({"message": "Defects updated", "allRoundsData": team_data["allRoundsData"]})


@app.route('/balls_in_play')
def balls_in_play():
    team_data = state["teams"][state["active_team"]]
    last_round = next((r for r in team_data["allRoundsData"] if r["round"] == state["current_round"]), None)
    balls_in_play = last_round.get("balls_in_play", 0) if last_round else 0
    return render_template('balls_in_play.html', current_round=state["current_round"], balls_in_play=balls_in_play,
                           active_team=state["active_team"])


@app.route('/update_balls_in_play', methods=['POST'])
def update_balls_in_play():
    balls_value = request.form.get('balls_in_play')
    if not balls_value or not balls_value.isdigit():
        return jsonify({"error": "Invalid balls value"}), 400

    balls_value = int(balls_value)
    team_data = state["teams"][state["active_team"]]
    for round_data in team_data["allRoundsData"]:
        if round_data["round"] == state["current_round"]:
            round_data["balls_in_play"] = balls_value
            break

    return jsonify({"message": "Balls in play updated"})


@app.route('/metric', methods=['GET'])
def metric():
    team_data = state["teams"][state["active_team"]]
    current_round = state["current_round"]

    # Get the round data
    round_data = next((r for r in team_data["allRoundsData"] if r["round"] == current_round), None)

    if round_data:
        total = round_data.get("total", 0)
        defects = round_data.get("defects", 0)
        balls_in_play = round_data.get("balls_in_play", 0)
        plan_value = round_data.get("plan", 0)

        print(f"[DEBUG] Saving to Excel: total={total}, defects={defects}, balls_in_play={balls_in_play}, plan={plan_value}")
        save_points_to_excel(total, defects, balls_in_play, plan_value)

        delta = (total - defects + 0.5 * balls_in_play) - plan_value
        num = 15*total
        if(abs(delta)==0):
            denom = 0.1
        else:
            denom=abs(delta)/9
        points = num/denom
    else:
        total = defects = balls_in_play = plan_value = points = 0

    return render_template(
        'metric.html',
        current_round=current_round,
        allRoundsData=team_data["allRoundsData"],
        active_team=state["active_team"],
        points=round(points, 2)
    )


@app.route('/update_round', methods=['POST'])
def update_round():
    plan_value = request.form.get('plan_value')
    final_count = request.form.get('final_count')

    if not plan_value or not plan_value.isdigit() or not final_count or not final_count.isdigit():
        return jsonify({"error": "Invalid data"}), 400

    plan_value = int(plan_value)
    final_count = int(final_count)

    team_data = state["teams"][state["active_team"]]
    for round_data in team_data["allRoundsData"]:
        if round_data["round"] == state["current_round"]:
            round_data["total"] = final_count
            break

    return jsonify({"message": "Round updated"})


@app.route('/strategize')
def strategize_page():
    reset_arduino()
    team_data = state["teams"][state["active_team"]]

    # Add balls_in_play = 0 if not set for any round (safe fallback)
    for rd in team_data["allRoundsData"]:
        if "balls_in_play" not in rd:
            rd["balls_in_play"] = 0

    return render_template(
        'strategize.html',
        current_round=state["current_round"],
        allRoundsData=team_data["allRoundsData"],
        active_team=state["active_team"],
        teams_enabled=state["teams_enabled"]
    )


@app.route('/next_round', methods=['POST'])
def next_round():
    plan_value = request.form.get('plan_value')
    if not plan_value or not plan_value.isdigit():
        return jsonify({"error": "Invalid plan value"}), 400

    plan_value = int(plan_value)
    team = state["active_team"]
    team_data = state["teams"][team]
    existing_round = next((r for r in team_data["allRoundsData"] if r["round"] == state["current_round"]), None)
    if not existing_round:
        team_data["allRoundsData"].append({
            "round": state["current_round"],
            "plan": plan_value,
            "total": team_data["balls_counter"],
            "defects": 0
        })

    if state["teams_enabled"]:
        if state["active_team"] == "Team A":
            state["active_team"] = "Team B"
        else:
            state["active_team"] = "Team A"
            state["current_round"] += 1
    else:
        state["current_round"] += 1

    return jsonify({"current_round": state["current_round"], "active_team": state["active_team"]})


@app.route('/clear_data', methods=['POST'])
def clear_data():
    state["current_round"] = 1
    for team in state["teams"]:
        state["teams"][team]["balls_counter"] = 0
        state["teams"][team]["allRoundsData"] = []
    state["active_team"] = "Team A"
    return jsonify({"message": "Data cleared", "current_round": state["current_round"]})


@app.route('/live_counter')
def live_counter():
    def stream():
        arduino = get_arduino_connection()
        if arduino:
            try:
                while True:
                    response = arduino.readline().decode().strip()
                    if response.isdigit():
                        state["teams"][state["active_team"]]["balls_counter"] = int(response)
                        yield f"data: {response}\n\n"
            except Exception as e:
                print(f"Error fetching live counter: {e}")
            finally:
                arduino.close()

    return Response(stream_with_context(stream()), mimetype="text/event-stream")


if __name__ == '__main__':
    import threading
    import time
    import webbrowser
    import os

    def start_server():
        app.run(port=5000, debug=False, use_reloader=False)

    server_thread = threading.Thread(target=start_server, daemon=True)
    server_thread.start()

    if os.environ.get("WERKZEUG_RUN_MAIN") == "true" or os.environ.get("WERKZEUG_RUN_MAIN") is None:
        time.sleep(2)
       # webbrowser.open("http://127.0.0.1:5000")

    # Keep the main thread alive
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("Server stopped.")
